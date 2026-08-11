#!/usr/bin/env python3
"""Generate the Red Pill sample apparel MIS workbook (Example + Template + Instructions).

Deterministic (seed=42) — re-running reproduces the identical workbook.
Run from anywhere:  python examples/generate_sample.py  (requires openpyxl)
Outputs (in cwd):   RedPill_Sample_MIS_Apparel.xlsx  +  sample_example.csv
"""
import csv, math, random
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

random.seed(42)
BF = 1.5  # buffer factor the engine uses

# ---------------------------------------------------------------- master data
STORES = [  # (name, store_factor, lead_time_base_days)
    ("Mumbai", 1.30, 4), ("Delhi", 1.25, 6), ("Bangalore", 1.20, 7),
    ("Hyderabad", 1.00, 7), ("Pune", 0.95, 5), ("Chennai", 0.90, 9),
    ("Kolkata", 0.80, 10), ("Ahmedabad", 0.75, 6),
]

# sku_code, style, category, color, size, unit_price, base_daily_demand, lt_add
SKUS = [
    ("TSH-CRW-BLK-M", "Crew Neck T-Shirt", "Tops", "Black", "M", 799, 9, 0),
    ("TSH-CRW-WHT-L", "Crew Neck T-Shirt", "Tops", "White", "L", 799, 8, 0),
    ("TSH-CRW-NVY-S", "Crew Neck T-Shirt", "Tops", "Navy", "S", 799, 6, 0),
    ("TSH-CRW-GRN-XL", "Crew Neck T-Shirt", "Tops", "Green", "XL", 799, 5, 0),
    ("POL-TSH-RED-M", "Polo T-Shirt", "Tops", "Red", "M", 1199, 6, 0),
    ("POL-TSH-GRN-L", "Polo T-Shirt", "Tops", "Green", "L", 1199, 5, 0),
    ("POL-TSH-NVY-S", "Polo T-Shirt", "Tops", "Navy", "S", 1199, 5, 0),
    ("JNS-SLM-IND-32", "Slim Fit Jeans", "Bottoms", "Indigo", "32", 2499, 5, 1),
    ("JNS-SLM-BLK-34", "Slim Fit Jeans", "Bottoms", "Black", "34", 2499, 4, 1),
    ("JNS-SLM-IND-30", "Slim Fit Jeans", "Bottoms", "Indigo", "30", 2499, 4, 1),
    ("CHN-TRS-KHK-32", "Chino Trousers", "Bottoms", "Khaki", "32", 1799, 4, 1),
    ("CHN-TRS-NVY-34", "Chino Trousers", "Bottoms", "Navy", "34", 1799, 3, 1),
    ("CHN-TRS-OLV-30", "Chino Trousers", "Bottoms", "Olive", "30", 1799, 3, 1),
    ("SHT-OXF-BLU-M", "Oxford Shirt", "Shirts", "Blue", "M", 1899, 4, 0),
    ("SHT-OXF-WHT-L", "Oxford Shirt", "Shirts", "White", "L", 1899, 4, 0),
    ("SWT-HOD-GRY-M", "Hooded Sweatshirt", "Outerwear", "Grey", "M", 2199, 4, 1),  # chronic-underestimated
    ("SWT-HOD-BLK-XL", "Hooded Sweatshirt", "Outerwear", "Black", "XL", 2199, 3, 1),  # chronic-underestimated
    ("JKT-DNM-BLU-M", "Denim Jacket", "Outerwear", "Blue", "M", 3499, 3, 2),  # volatile
    ("JKT-DNM-BLU-L", "Denim Jacket", "Outerwear", "Blue", "L", 3499, 2, 2),  # volatile
    ("SHR-RUN-BLK-M", "Running Shorts", "Activewear", "Black", "M", 999, 6, 0),
    ("SHR-RUN-BLU-L", "Running Shorts", "Activewear", "Blue", "L", 999, 5, 0),
    ("KUR-ETH-MRN-M", "Kurta", "Ethnic", "Maroon", "M", 1599, 4, 1),
    ("KUR-ETH-CRM-L", "Kurta", "Ethnic", "Cream", "L", 1599, 3, 1),
    ("BLZ-FRM-CHR-40", "Formal Blazer", "Formal", "Charcoal", "40", 5999, 2, 2),
    ("BLZ-FRM-NVY-42", "Formal Blazer", "Formal", "Navy", "42", 5999, 1, 2),
]

CHRONIC = {"SWT-HOD-GRY-M", "SWT-HOD-BLK-XL"}   # stated ADS too low, rising sales
VOLATILE = {"JKT-DNM-BLU-M", "JKT-DNM-BLU-L"}   # high week-to-week variance
HERO = {"TSH-CRW-BLK-M", "JNS-SLM-IND-32", "SHR-RUN-BLK-M", "POL-TSH-RED-M", "KUR-ETH-MRN-M"}


def engine_status(soh, qoo, ads, lt):
    pipeline = soh + qoo
    rop = ads * lt
    buf = ads * lt * BF
    if soh == 0 and qoo == 0: return "OUT_OF_STOCK"
    if soh == 0 and qoo > 0:  return "INCOMING"
    if pipeline < 0.5 * rop:  return "CRITICAL"
    if pipeline < rop:        return "REORDER"
    if soh > 2 * buf:         return "OVERSTOCK"
    return "OPTIMAL"


def set_stock(status, ads, lt):
    rop = ads * lt
    soh = qoo = 0
    for _ in range(60):
        if status == "OUT_OF_STOCK":  soh, qoo = 0, 0
        elif status == "INCOMING":    soh, qoo = 0, max(1, round(rop * random.uniform(0.7, 1.3)))
        elif status == "CRITICAL":    soh, qoo = max(1, round(rop * random.uniform(0.15, 0.42))), 0
        elif status == "REORDER":     soh, qoo = max(1, round(rop * random.uniform(0.55, 0.9))), 0
        elif status == "OVERSTOCK":
            tot = round(3 * ads * lt * random.uniform(1.2, 1.9)) + 3
            qoo = round(tot * random.choice([0, 0, 0.2])); soh = tot - qoo
        else:  # OPTIMAL
            tot = round(ads * lt * random.uniform(1.2, 2.6))
            qoo = round(tot * random.choice([0, 0, 0, 0.3])); soh = max(1, tot - qoo)
        if engine_status(soh, qoo, ads, lt) == status:
            return soh, qoo
    return soh, qoo  # best effort


def weekly_sales(ads_actual, mode):
    mu = ads_actual * 7.0
    if mode == "volatile":
        vals = [max(0, round(random.gauss(mu, mu * 0.7))) for _ in range(8)]
    elif mode == "uptrend":
        vals = [max(0, round(mu * f * random.uniform(0.9, 1.1)))
                for f in (0.6, 0.7, 0.8, 0.95, 1.1, 1.25, 1.4, 1.55)]
    else:
        vals = [max(0, round(random.gauss(mu, mu * 0.18))) for _ in range(8)]
    return vals


POOL = (["OPTIMAL"] * 44 + ["REORDER"] * 17 + ["CRITICAL"] * 9 +
        ["OVERSTOCK"] * 12 + ["OUT_OF_STOCK"] * 6 + ["INCOMING"] * 5 + ["OPTIMAL"] * 7)

rows = []
for si, (code, style, cat, color, size, price, base, lt_add) in enumerate(SKUS):
    for ti, (store, sfac, lt_base) in enumerate(STORES):
        ads_actual = max(1, round(base * sfac))
        lt = max(2, lt_base + lt_add + random.choice([-1, 0, 0, 1]))
        mode = "normal"; stated = ads_actual
        if code in CHRONIC:
            mode = "uptrend"; stated = max(1, round(ads_actual * 0.6))
        elif code in VOLATILE:
            mode = "volatile"; stated = ads_actual
        else:
            r = random.random()
            if r < 0.12:   stated = max(1, round(ads_actual * 0.65))
            elif r < 0.20: stated = max(1, round(ads_actual * 1.5))
        sales = weekly_sales(ads_actual, mode)
        if code in HERO and ti == 0:      status = "OVERSTOCK"
        elif code in HERO and ti == 1:    status = random.choice(["OUT_OF_STOCK", "CRITICAL"])
        elif code in CHRONIC:             status = random.choice(["REORDER", "REORDER", "CRITICAL", "OPTIMAL"])
        else:                             status = random.choice(POOL)
        soh, qoo = set_stock(status, stated, lt)
        rows.append({
            "sku_code": code, "style": style, "category": cat, "color": color, "size": size,
            "store": store, "soh": soh, "qoo": qoo, "ads": stated, "lead_time": lt,
            "unit_price": price,
            "sold_w1": sales[0], "sold_w2": sales[1], "sold_w3": sales[2], "sold_w4": sales[3],
            "sold_w5": sales[4], "sold_w6": sales[5], "sold_w7": sales[6], "sold_w8": sales[7],
        })

def find(code, store):
    return next(r for r in rows if r["sku_code"] == code and r["store"] == store)

find("CHN-TRS-NVY-34", "Chennai")["ads"] = ""       # blank ADS
find("SHT-OXF-BLU-M", "Kolkata")["lead_time"] = ""  # missing lead time
find("KUR-ETH-CRM-L", "Ahmedabad")["soh"] = "N/A"   # non-numeric stock

COLS = ["sku_code", "style", "category", "color", "size", "store", "soh", "qoo",
        "ads", "lead_time", "unit_price",
        "sold_w1", "sold_w2", "sold_w3", "sold_w4", "sold_w5", "sold_w6", "sold_w7", "sold_w8"]

with open("sample_example.csv", "w", newline="") as f:
    w = csv.DictWriter(f, fieldnames=COLS); w.writeheader()
    for r in rows: w.writerow(r)

HEADERS = {"sku_code": "SKU Code", "style": "Style", "category": "Category", "color": "Color",
           "size": "Size", "store": "Store", "soh": "SOH", "qoo": "QOO", "ads": "ADS",
           "lead_time": "Lead Time", "unit_price": "Unit Price", "sold_w1": "Sold W1",
           "sold_w2": "Sold W2", "sold_w3": "Sold W3", "sold_w4": "Sold W4", "sold_w5": "Sold W5",
           "sold_w6": "Sold W6", "sold_w7": "Sold W7", "sold_w8": "Sold W8"}
FONT = "Arial"; RED = "C0392B"; DARK = "1A1A1A"; LGRAY = "F2F2F2"; YELL = "FFF2CC"
thin = Side(style="thin", color="D9D9D9"); border = Border(left=thin, right=thin, top=thin, bottom=thin)
wb = openpyxl.Workbook()

def style_header(ws, ncols, rownum=1, fill=RED):
    for c in range(1, ncols + 1):
        cell = ws.cell(row=rownum, column=c)
        cell.font = Font(name=FONT, bold=True, color="FFFFFF", size=10)
        cell.fill = PatternFill("solid", fgColor=fill)
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = border

def write_data_sheet(ws, data_rows, example_only=False):
    ws.sheet_view.showGridLines = False
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=len(COLS))
    t = ws.cell(row=1, column=1, value=("RED PILL — Weekly MIS  ·  EXAMPLE (dummy data)" if not example_only
                                        else "RED PILL — Weekly MIS  ·  TEMPLATE (fill this in)"))
    t.font = Font(name=FONT, bold=True, size=13, color=RED)
    t.alignment = Alignment(horizontal="left", vertical="center"); ws.row_dimensions[1].height = 24
    for i, key in enumerate(COLS, start=1): ws.cell(row=2, column=i, value=HEADERS[key])
    style_header(ws, len(COLS), rownum=2); ws.row_dimensions[2].height = 30
    for ri, r in enumerate(data_rows, start=3):
        for ci, key in enumerate(COLS, start=1):
            cell = ws.cell(row=ri, column=ci, value=r.get(key, "")); cell.font = Font(name=FONT, size=10)
            cell.border = border
            if key == "sku_code": cell.font = Font(name=FONT, size=10, bold=True)
            if key in ("soh", "qoo", "ads", "lead_time") or key.startswith("sold_"):
                cell.alignment = Alignment(horizontal="center")
            if key == "unit_price":
                cell.number_format = '₹#,##0'; cell.alignment = Alignment(horizontal="right")
            if ri % 2 == 1: cell.fill = PatternFill("solid", fgColor=LGRAY)
    ws.freeze_panes = "G3"
    widths = {"sku_code": 16, "style": 18, "category": 11, "color": 9, "size": 6, "store": 12,
              "soh": 7, "qoo": 7, "ads": 7, "lead_time": 10, "unit_price": 11}
    for i, key in enumerate(COLS, start=1):
        ws.column_dimensions[get_column_letter(i)].width = widths.get(key, 8.5)

ws1 = wb.active; ws1.title = "① Example"; write_data_sheet(ws1, rows)
ws2 = wb.create_sheet("② Template")
example_row = {"sku_code": "TSH-CRW-BLK-M", "style": "Crew Neck T-Shirt", "category": "Tops",
               "color": "Black", "size": "M", "store": "Mumbai", "soh": 8, "qoo": 0, "ads": 12,
               "lead_time": 4, "unit_price": 799, "sold_w1": 70, "sold_w2": 82, "sold_w3": 78,
               "sold_w4": 85, "sold_w5": 90, "sold_w6": 88, "sold_w7": 95, "sold_w8": 92}
write_data_sheet(ws2, [example_row], example_only=True)
for c in range(1, len(COLS) + 1): ws2.cell(row=3, column=c).fill = PatternFill("solid", fgColor=YELL)
ws2.cell(row=5, column=1, value="↑ Example row — DELETE it, then add ONE row per SKU × Store. "
         "Required: SKU Code, Store, SOH, ADS, Lead Time. QOO blank = 0. "
         "Sold W1..W8 (units sold per week; W8 = most recent) are optional but power ADS auto-correction.")
ws2.cell(row=5, column=1).font = Font(name=FONT, italic=True, size=9, color="7F7F7F")
ws2.merge_cells(start_row=5, start_column=1, end_row=5, end_column=len(COLS))

ws3 = wb.create_sheet("③ Instructions"); ws3.sheet_view.showGridLines = False
ws3.column_dimensions["A"].width = 20; ws3.column_dimensions["B"].width = 14; ws3.column_dimensions["C"].width = 70
def h(row, text, size=13, color=RED):
    c = ws3.cell(row=row, column=1, value=text); c.font = Font(name=FONT, bold=True, size=size, color=color)
r = 1; h(r, "Red Pill — How to use this workbook"); r += 1
ws3.cell(row=r, column=1, value="Right SKU, right place, right time — Theory-of-Constraints inventory analysis.")
ws3.cell(row=r, column=1).font = Font(name=FONT, italic=True, size=10, color="7F7F7F"); r += 2
h(r, "Run it in Claude", size=11); r += 1
for line in ["1. Open Claude Code (or the Claude app) with the Red Pill skill installed.",
             "     /plugin marketplace add koushikeverything/redpill",
             "     /plugin install redpill@koushik-skills",
             "2. Fill the '② Template' sheet with your data (one row per SKU × Store).",
             "3. Attach this file and say:  “run red pill on this weekly MIS report”.",
             "4. Claude returns a health report + demand plan (orders, transfers, ADS corrections)."]:
    ws3.cell(row=r, column=1, value=line).font = Font(name=FONT, size=10)
    ws3.merge_cells(start_row=r, start_column=1, end_row=r, end_column=3); r += 1
r += 1
h(r, "Columns", size=11); r += 1
ws3.cell(row=r, column=1, value="Column"); ws3.cell(row=r, column=2, value="Required?"); ws3.cell(row=r, column=3, value="Meaning")
style_header(ws3, 3, rownum=r, fill=DARK); r += 1
for name, req, mean in [
    ("SKU Code", "Yes", "Unique item/variant identifier (style-color-size). Transfers pair by this."),
    ("Style / Category / Color / Size", "Optional", "Descriptive attributes; pass through untouched."),
    ("Store", "Yes", "Selling location. Same SKU can differ by store."),
    ("SOH", "Yes", "Stock on hand — units physically in the store now (>= 0)."),
    ("QOO", "No", "Quantity on order / in transit. Blank = 0 (assumed, and disclosed)."),
    ("ADS", "Yes", "Average daily sales (units/day) for THIS SKU at THIS store — your master value."),
    ("Lead Time", "Yes", "Supplier lead time in DAYS for this SKU at this store (> 0). Never assumed."),
    ("Unit Price", "Optional", "Per-unit price (₹). Enables order values and transfer savings."),
    ("Sold W1..W8", "Optional", "Units sold per week; W8 = most recent, W1 = 8 weeks ago. Lets Claude auto-correct ADS.")]:
    ws3.cell(row=r, column=1, value=name).font = Font(name=FONT, size=10, bold=True)
    rc = ws3.cell(row=r, column=2, value=req); rc.font = Font(name=FONT, size=10, bold=(req == "Yes"),
                  color=(RED if req == "Yes" else "7F7F7F")); rc.alignment = Alignment(horizontal="center")
    ws3.cell(row=r, column=3, value=mean).font = Font(name=FONT, size=10)
    for c in range(1, 4): ws3.cell(row=r, column=c).border = border
    r += 1
r += 1
h(r, "Status legend (what Claude computes)", size=11); r += 1
for name, mean in [("⚫ OUT OF STOCK", "SOH = 0 and nothing on order — losing sales now."),
                   ("🟣 INCOMING", "SOH = 0 but replenishment in transit."),
                   ("🔴 CRITICAL", "Will stock out before the next delivery can land."),
                   ("🟡 REORDER", "Order now; still time if you act today."),
                   ("🔵 OVERSTOCK", "Capital tied up — a donor for inter-store transfers."),
                   ("🟢 OPTIMAL", "Buffer healthy; no action needed.")]:
    ws3.cell(row=r, column=1, value=name).font = Font(name=FONT, size=10, bold=True)
    ws3.cell(row=r, column=3, value=mean).font = Font(name=FONT, size=10); r += 1

wb.save("RedPill_Sample_MIS_Apparel.xlsx")
print("wrote RedPill_Sample_MIS_Apparel.xlsx and sample_example.csv  |  rows:", len(rows))
